import openpyxl as op
from openpyxl import Workbook, load_workbook
import requests, os, rasterio, rasterio.features, rasterio.warp
from datetime import datetime
from pprint import pprint
import string
import re
#FUNCION PARA CONSEGUIR LAT/LONG DEL GEOTIFF Y OBTENER DATA
     #La variable de entrada es un tif
def OpenGeoTiffAndGetData (path):
    with rasterio.open(path) as dataset:

        # Read the dataset's valid data mask as a ndarray.
        mask = dataset.dataset_mask()

        # Extract feature shapes and values from the array.
        for geom, val in rasterio.features.shapes(
                mask, transform=dataset.transform):

            # Transform shapes from the dataset's own coordinate
            # reference system to CRS84 (EPSG:4326).
            geom = rasterio.warp.transform_geom(
                dataset.crs, 'EPSG:4326', geom, precision=6)

            # Print GeoJSON shapes to stdout.
            longitude=(geom['coordinates'][0][0][0])
            latitude=(geom['coordinates'][0][0][1])
            #print('Latitud:', latitude)
            #print('Longitud:', longitude)
    
    return longitude, latitude
Lon, Lat =OpenGeoTiffAndGetData('/home/pi/Desktop/8-GIT/repos/resources/NDVI_Olivos_MALAGA.tif')

def ReadingExcel (path, Longitud, Latitud):
    wb = op.load_workbook(path)
    ws = wb.active
    d=datetime.today()
    mes=d.month
    #mes = int(input("Introduzca el mes a consultar en formato numerico para las horas de insolación: "))
    #HEMISFERIO NORTE
    NHemis = True
    if Latitud < 0:
        print("Hemisferio sur")
        mes = mes+13
        NHemis = False
        Latitud = -Latitud
    isEven = False
    #print("Latitud en valor absoluto ", Latitud)
    LatitudOriginal = Latitud
    Latitud = round(Latitud)
        
    temp = Latitud%2
    if temp == 0:
        isEven=True
    if True!=isEven:
        if(LatitudOriginal-Latitud > 0):
            Latitud = Latitud +1
        else:
            Latitud=Latitud-1
    #print("Latitud sin decimales  formateada para tabla " , Latitud)
    ActualRow = int(38-Latitud/2)
    result = ""
    for row in ws.iter_rows(min_row=ActualRow, max_col=mes, min_col=mes, max_row=ActualRow, values_only=True):
        #print(row[0])
        result = row[0]
        n=result
    return n
n=ReadingExcel('/home/pi/Desktop/8-GIT/repos/documentation/22-Insolacion_zonas.xlsx', Lon, Lat)
